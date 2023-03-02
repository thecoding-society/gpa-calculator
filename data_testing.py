from openpyxl import Workbook, load_workbook
from program import *


wb = load_workbook(filename = 'test1.xlsx')

sheet = wb['Sheet1']



marks = []
credits = [sheet.cell(row = 1, column = i).value for i in range(1, 9)]



for student in range(2, len(sheet['A']) + 1):
    gpa = Gpa()

    for i in range(1, 9):
        gpa.addsubject(sheet.cell(row = 1, column = i).value, sheet.cell(row = student, column = i).value)

        marks.append(sheet.cell(row = student, column = i).value)

    gpa.calc()
    gpa.display()

    sheet.cell(row = student, column = 10).value = gpa.gpa

    wb.save("done.xlsx")
    del gpa

